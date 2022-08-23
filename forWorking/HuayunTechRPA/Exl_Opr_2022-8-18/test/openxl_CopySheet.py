#  Author : Github: @GWillS163
#  Time: $(Date)

from openpyxl import load_workbook  # 读取时导入这个
from openpyxl.styles import Font, Alignment  # 设置单元格格式
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

file = r"D:\Project\python_scripts\forWorking\HuayunTechRPA\Exl_Opr_2022-8-18\test.xlsx"
wb = load_workbook(file)
ws = wb["测试问卷"]

new = wb.create_sheet("测试问卷_copy", index=0)

# copy ws A1:K10 to new sheet
for row in ws["A1":"K10"]:
    for cell in row:
        new[cell.coordinate].value = cell.value
        new[cell.coordinate].font = copy(cell.font)
        new[cell.coordinate].alignment = copy(cell.alignment)
        new[cell.coordinate].border = copy(cell.border)
        new[cell.coordinate].fill = copy(cell.fill)
        new[cell.coordinate].number_format = copy(cell.number_format)
        new[cell.coordinate].protection = copy(cell.protection)
        new[cell.coordinate].style = copy(cell.protection)

        # new[cell.coordinate].style = copy(cell.style)
# new.cell(1,1).value = ws.cell(1,1).value
# new.cell(1,1).font = copy(ws.cell(1,1).font)
# new.cell(1,1).alignment = copy(ws.cell(1,1).alignment)
# new.cell(1,1).fill = copy(ws.cell(1,1).fill)
# new.cell(1,1).border = copy(ws.cell(1,1).border)
# new.cell(1,1).protection = copy(ws.cell(1,1).protection)
# new.cell(1,1).number_format = copy(ws.cell(1,1).number_format)

wb.save("copyTest.xlsx")
